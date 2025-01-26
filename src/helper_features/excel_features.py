# pylint: disable=missing-docstring

from pathlib import Path
import logging
import openpyxl
import pandas as pd
import os
from src.data_handlers import Globals

logger = logging.getLogger(__name__)

# pylint: disable=line-too-long


class ExcelOperations:
    """Class to perform Excel functions"""

    def __init__(self):
        self.excel_file_folder = self.determine_excel_file_folder()


    def determine_excel_file_folder(self) -> str:
        """Function to determine the folder of the Excel file"""
        if Globals.ANALYSIS_TYPE == "cli":
            return "data_cli_analysis"
        if Globals.ANALYSIS_TYPE == "full":
            return "data_full_analysis"
        if Globals.ANALYSIS_TYPE == "earnings_calendar":
            return "data_earnings_analysis"



    # # @staticmethod
    # def get_abs_path(self, folder_name: str) -> Path:
    #     """Function to get the absolute path of the directory
    #     :return: Absolute path of the directory
    #     """
    #     current_path = Path.cwd()
    #     path_ending = "api_stock_from_scratch"
    #     absolute_path = None
    #     # Traverse through parent folders
    #
    #     for parent in current_path.parents:
    #         if str(parent).endswith(path_ending):
    #             absolute_path = Path.joinpath(parent, folder_name)
    #             break
    #
    #     if absolute_path is None:
    #         logger.warning("Path not found")
    #     return absolute_path



    async def clear_sheets(self):
        """Obtains and clear all sheets for Excel files in the directory."""
        # get file path

        # folder_path = ExcelOperations.get_abs_path(self)

        excel_folder = find_folder(self.excel_file_folder)
        path_object = Path(excel_folder)
        print(f"excel folder: {excel_folder}")

        # print(f"folder path: {folder_path}")

        # Iterate through the files in the directory

        for file_path in path_object.iterdir():
            # Check if it's a file (not a subdirectory) and not a recovery file

            if file_path.is_file() and file_path.stem[0].isalpha():
                ExcelOperations.execute_clearing_sheets(file_path=file_path)

    @staticmethod
    def execute_clearing_sheets(file_path: Path):
        """Function to clear the sheets in the Excel file
        :param file_path: Path to the Excel file
        """
        # load excel file

        book = openpyxl.load_workbook(file_path)
        # iterate through the sheets in the file

        for sheet in book.sheetnames:
            current_sheet = book[sheet]
            # if sheet not in list_df_names and sheet != "Sheet1":

            if sheet != "Sheet1":
                print(f"deleting sheet {sheet}")
                del book[sheet]
                continue
            current_sheet.delete_rows(2, current_sheet.max_row - 1)
        # save the file to the path

        book.save(file_path)

    def get_abs_path(self) -> Path:
        """Function to get the absolute path of the directory to write the Excel files
        :return: Absolute path of the directory
        """
        current_path = Path.cwd()

        # if the analysis is done suing the cli purposes, then move up two directories to the current directory
        if self.excel_file_folder == "data_cli_analysis":
            grandparent_path = current_path.parent.parent
            abs_path = Path.joinpath(grandparent_path, self.excel_file_folder)

            return abs_path
        abs_path = Path.joinpath(current_path, self.excel_file_folder)

        return abs_path


    async def load_dataframes_into_excel(self):
        """Function to load dataframes into excel files"""
        # get file path

        # folder_path = self.get_abs_path()
        excel_folder = find_folder(self.excel_file_folder)
        path_object = Path(excel_folder)

        # Iterate through the files in the directory

        for file_path in path_object.iterdir():
            # Check if it's a file (not a subdirectory)

            if file_path.is_file():
                # add dataframes to the corresponding excel file

                # if the analysis is done for testing or cli purposes, then write the share to the general dataframe

                if "general" in str(file_path) and Globals.ANALYSIS_TYPE == "cli":
                    await self.write_to_excel(
                        list_df=Globals.DATAFRAME_LIST_GENERAL, file_path=file_path
                    )
                if "dip" in str(file_path):
                    await self.write_to_excel(
                        list_df=Globals.DATAFRAME_LIST_DIP_STOCKS, file_path=file_path
                    )
                if "undervalued" in str(file_path):
                    await self.write_to_excel(
                        list_df=Globals.DATAFRAME_LIST_UNDERVALUED_STOCKS,
                        file_path=file_path,
                    )
                if "growth" in str(file_path):
                    await self.write_to_excel(
                        list_df=Globals.DATAFRAME_LIST_GROWTH_STOCKS,
                        file_path=file_path,
                    )
                if "magic" in str(file_path):
                    await self.write_to_excel(
                        list_df=Globals.DATAFRAME_LIST_MAGIC_FORMULA,
                        file_path=file_path,
                    )

    @staticmethod
    async def write_to_excel(list_df: list, file_path: Path):
        """Function to write dataframes to Excel file
        :param list_df: List of dataframes to write
        :param file_path: Path to the Excel file
        """
        for dataframe in list_df:
            with pd.ExcelWriter(  # pylint: disable=abstract-class-instantiated
                    path=file_path,
                    engine="openpyxl",
                    mode="a",
                    if_sheet_exists="replace",
            ) as writer:
                dataframe.to_excel(writer, sheet_name=dataframe.attrs["name"])



def find_folder(folder_name, search_path="C:\\Users\\m.hofman\\Documents\\python_projects\\api_stock_from_scratch"):
    for root, dirs, files in os.walk(search_path):
        if folder_name in dirs:
            return os.path.join(root, folder_name)
    return None